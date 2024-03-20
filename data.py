
from openpyxl.reader.excel import load_workbook

class WebData:
    """this class is used to contain all the data required to test OrangeHRM
    data is stored in the excel file"""

    def __init__(self):
        self.url = "https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"
        self.dashboardURL = "https://opensource-demo.orangehrmlive.com/web/index.php/dashboard/index"
        self.fileName = "Data/test_data.xlsx"
        self.sheetName = "Sheet1"
        self.workbook = load_workbook(self.fileName)
        self.sheet = self.workbook[self.sheetName]

    def rowCount(self):
        """
        This method returns the maximum number of rows present in the Sheet 1
        :return: int
        """
        return self.sheet.max_row

    def readData(self, row, column):
        """this method returns the data present in the particular cell"""
        return self.sheet.cell(row, column).value

    def writeData(self, row, column, data):
        """this method id to write data in the excel sheet and save it"""
        self.sheet.cell(row, column).value = data
        self.workbook.save(self.fileName)

    def writeTestResult(self, row, column,  test_result):
        """this method is to write the test result in the excel sheet"""
        self.writeData(row, 7, test_result)

    def writeTestDate(self, row, column, date):
        """this method is to write the test date in the excel sheet"""
        self.writeData(row, 4, date)

    def writeTestTime(self, row, column, time):
        """this method is to write the test time in the excel sheet"""
        self.writeData(row, 5, time)

